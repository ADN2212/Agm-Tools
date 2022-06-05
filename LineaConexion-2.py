import tkinter as tk
from tkinter import messagebox
from math import atan
from math import pi
from math import hypot
from io import open
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import colors, Font, Border, Side, Alignment


ventana = tk.Tk()
ventana.geometry("210x180")
ventana.title("LC_V-2")
ventana.resizable(False, False)
#---------------------------------------------------------------------------Variables-----------------------------------------------------------------------------------------------------------

puntos = []


#--------------------------------------------------------------------Funciones y Clase----------------------------------------------------------------------------------------------------------
class Punto:	
	def __init__(self, nombre, este, norte):
		self.nombre = nombre
		self.este = este
		self.norte = norte

def calcular_GMS(angulo, c):
	"""
	Esta función recibe el angulo en grados sexagesimales decimales.
	El parametro c indaca hasta donde llegara el angulo, 1 = solo grados, 2 = grados y minutos, 3 = grados, minutos y segundos.
	"""
	if angulo < 0:
		angulo = -1 * angulo

	entero = 0
	decimal = 0
	GMS = [0, 0, 0]

	for i in range(3):
		entero = int(angulo)
		decimal = angulo - entero

		if entero < 10:
			GMS[i] = "0" + str(entero)#Para agragar el cero delante, esto es con fines esteticos.
		else:
			GMS[i] = entero

		angulo = decimal * 60

	if c == 3:	
		return str(GMS[0]) + "°" + str(GMS[1]) + "'" + str(GMS[2]) + '"'

	elif c == 2:
		return str(GMS[0]) + "°" + str(GMS[1]) + "'" 
	
	elif c == 1:
		return str(GMS[0]) + "°"

	else:
		return None		



def rumbo_y_dist(punto_1, punto_2):
	"""Esta funcion recibe dos objetos de tipo Punto para retornar una tupla con la distancia y el rumbo"""

	delta_x = punto_2.este - punto_1.este
	delta_y = punto_2.norte - punto_1.norte
	distancia = round(hypot(delta_x, delta_y), 2)
	rumbo = "No definido"
#Es bueno enlazarlos para que una vez se cumpla una ignore las demas

	if delta_x > 0 and delta_y > 0:
		rumbo = "N {} E".format(calcular_GMS(atan(delta_x/delta_y)*(180/pi), 2))

	elif delta_x > 0 and delta_y < 0:
		rumbo = "S {} E".format(calcular_GMS(atan(delta_x/delta_y)*(180/pi), 2))

	elif delta_y < 0 and delta_y < 0:
		rumbo = "S {} W".format(calcular_GMS(atan(delta_x/delta_y)*(180/pi), 2))	

	elif delta_x < 0 and delta_y > 0:
		rumbo = "N {} W".format(calcular_GMS(atan(delta_x/delta_y)*(180/pi), 2))

	#Los siguientes casos son menos comunes:

	elif delta_x > 0 and delta_y == 0:
		rumbo = "N 90°00' E"#tambien puede ser "S 90°00' E"
	
	elif delta_x == 0 and delta_y < 0:
		rumbo = "S 00°00' E"#tambien puede ser "S 00° 00' W"

	elif delta_x < 0 and delta_y == 0:
		rumbo = "N 90°00' W"#tambien puede ser "S 90° 00' W"
	
	elif delta_x == 0 and delta_y > 0:
		rumbo = "N 00°00' E"#tambien puede ser "N 00° 00' W"

	return (rumbo, round(distancia, 2))

def mostrar_instrucciones():
	texto = "Introduzca un archivo de texto con los puntos que correspondan a las lineas de conexion, con el formato siguiente: \n\n"	
	texto2 = "Numero_Punto,Coordenada_Este,Coordenada_Norte \n\n"
	texto3 = "La información debe estar ordenada en el siguiente orden: PG-Visual, PG-Base o Pivote y demas puntos, pulse el botón 'Exportar Ejemplo' para ver un ejemplo."

	messagebox.showinfo("Instrucciones", texto + texto2 + texto3)

def probar_formato(cadena_text, num_linea):
	cumple = True

	if cadena_text.count(",") != 2:
		messagebox.showwarning("Error :(", "Error en la linea No.{}, no cumple con el formato establecido, sobran o faltan comas".format(num_linea))
		cumple = False
	else:
		cadena_text = cadena_text.split(",")
		try:
			Este = float(cadena_text[1])
			Norte = float(cadena_text[2])
			
			if Este < 0 or Norte < 0:
				messagebox.showwarning("Error", "Error en la linea No.{}, las coordenadas UTM no son negativas".format(num_linea))
				cumple = False

		except ValueError:
			messagebox.showwarning("Error", "Error de escritura en las coordenadas de la linea No.{}".format(num_linea))
			cumple = False
	
	return cumple		

def cargar():
	global puntos

	if len(puntos) != 0:
		puntos = []#En caso se que el usuario vuelva a cargar informacion, se borrara la que estaba con anterioridad.

	archivo = filedialog.askopenfilename(title="Abrir", initialdir = "desktop", filetypes = (("Ficheros de Texto", "*.txt"), ("Todos los Ficheros", "*.*")))	
	
	if archivo == '':#Para evitar que salte un error si el usuario decide cerrar la interfas de seleccion de archivo.
		pass
	
	else:
		archivo = open(archivo, "r")	
		archivo = archivo.readlines()
		pasaron = True

		while archivo[len(archivo) -1] == "\n":
			archivo.pop()#en caso de que el ultimo elemento sea un salto de linea, lo eliminará.
			
		for i in range(len(archivo)):			
			if not probar_formato(archivo[i], i + 1):
				pasaron = False

		if pasaron:#Esta linea sole se ejecutara en caso de que todos los puntos cumplan con el formato.
			for i in range(len(archivo)):
				archivo[i] = archivo[i].split(",")
				puntos.append(Punto(archivo[i][0], float(archivo[i][1]), float(archivo[i][2])))

			messagebox.showinfo(":)", "La información se a cargado con exito!!")

def dar_formato_tabla(hoja):
	#Configuraciones:
	fuente = Font(name = "Arial", size = 11)
	linea_borde = Side(border_style = "thin")
	alineacion = Alignment(horizontal = "center")
	todos_los_bordes = Border(top = linea_borde, left = linea_borde, right = linea_borde, bottom = linea_borde)
	formato_millares = '#,##0.00'
	maxima_fila = 4 + len(puntos)

	for fila in hoja.iter_rows(min_row = 3, max_row = maxima_fila, min_col = 2, max_col = 6):
		for celda in fila:
			celda.font = fuente
			celda.alignment = alineacion
			celda.border = todos_los_bordes

	#Aplicar el formato millares a las celdas que les corresponden:

	for fila in hoja.iter_rows(min_row = 5, max_row = maxima_fila, min_col = 4, max_col = 6):
		for celda in fila:
			celda.number_format = formato_millares
	
	#Las primeras dos filas llevan negritas:

	for fila in hoja.iter_rows(min_row = 3, max_row = 4, min_col = 2, max_col = 6):
		for celda in fila:
			celda.font = Font(name = "Arial", size = 11, bold = True)

	#Calcular los anchos de cada columna(Puede que esta parte no sea necesaria):

	ancho_B = 0
	ancho_C = 0
	ancho_D = 0
	ancho_E = 0
	ancho_F = 0

	for fila in hoja.iter_rows(min_row = 4, max_row = maxima_fila, min_col = 2, max_col = 6):
		
		largo = len(fila[0].value)
		if largo > ancho_B:
			ancho_B = largo

		largo = len(fila[1].value)
		if largo > ancho_C:
			ancho_C = largo

		largo = len(str(fila[2].value))
		if largo > ancho_D:
			ancho_D = largo

		largo = len(str(fila[3].value))			
		if largo > ancho_E:
			ancho_E = largo

		largo = len(str(fila[4].value))
		if largo > ancho_F:
			ancho_F = largo

	hoja.column_dimensions["B"].width = ancho_B + 5#El ancho de la cadena de texto mas grande no es suficiente.
	hoja.column_dimensions["C"].width = ancho_C + 5	
	hoja.column_dimensions["D"].width = ancho_D + 5
	hoja.column_dimensions["E"].width = ancho_E + 5
	hoja.column_dimensions["F"].width = ancho_F + 5

	hoja.merge_cells("B1:F1")
	hoja["B1"] = "Por Agrim Juan A. Núñez"
	hoja["B1"].font = Font(name = "times new roman", size = 12, bold = True, color = "000000FF")
	hoja["B1"].alignment = alineacion


def insertar_informacion():
	global puntos

	if len(puntos) == 0:
		messagebox.showinfo("Hey!!", "Todavia no has cargado ninguna información")
	
	else:
		
		direccion = filedialog.asksaveasfilename(defaultextension = ".xlsx", filetypes = (("Libro de Excel", "*.xlsx"), ("Todos los Ficheros", "*.*")))
		
		if direccion == '':#En caso de que el usuario cierre la interfas sin guardar el archivo.
			pass		
		
		else:
			#Crear libro y hoja:
			libro = Workbook()
			hoja = libro.active
			hoja.title = "From Python"

			#Encabezados:
			hoja.merge_cells("B3:F3")

			hoja["B3"] = "Lineas de Conexión"
			hoja["B4"] = "Estación"
			hoja["C4"] = "Rumbo"
			hoja["D4"] = "Distancia(mts)"
			hoja["E4"] = "Este(x)"
			hoja["F4"] = "Norte(y)"

			#Primeras dos lineas:

			ryd = rumbo_y_dist(puntos[0], puntos[1])
			hoja["B5"] = "{}-->{}".format(puntos[0].nombre, puntos[1].nombre)
			hoja["C5"] = ryd[0]
			hoja["D5"] = ryd[1]
			hoja["E5"] = round(puntos[0].este, 2)
			hoja["F5"] = round(puntos[0].norte, 2)

			hoja["B6"] = puntos[1].nombre
			hoja["C6"] = "---"
			hoja["D6"] = "---"
			hoja["E6"] = round(puntos[1].este, 2)
			hoja["F6"] = round(puntos[1].norte, 2)

			#Generar las demas lineas de conexion:
			
			i = 2#Empieza a iterar desde el tercer punto.

			for fila in hoja.iter_rows(min_row = 7, max_row = 7 + len(puntos)-3, min_col = 2, max_col = 6):
				ryd = rumbo_y_dist(puntos[1], puntos[i])
				fila[0].value = "{}-->{}".format(puntos[1].nombre, puntos[i].nombre)
				fila[1].value = ryd[0]
				fila[2].value = ryd[1]
				fila[3].value = round(puntos[i].este, 2)
				fila[4].value = round(puntos[i].norte, 2)
				i += 1

			dar_formato_tabla(hoja)	
			libro.save(direccion)
			messagebox.showinfo("Hecho", "la información ha sido guardada exitosamente en {}".format(direccion))


def exportar_ejemplo():
	direccion = filedialog.asksaveasfilename(defaultextension = ".txt", filetypes = (("Ficheros de Texto", "*.txt"), ("Todos los Ficheros", "*.*")))
	
	if direccion == '':
		pass

	else:
		ej = open(direccion, "w")
		texto = "PG-02,324769.85,2149715.17 --> Punto Visual\n"
		texto += "PG-01,324846.48,2149704.37 --> Punto Base o Pivote, desde el cual se conectan todos los puntos\n"
		texto += "E-1,324811.10,2149715.99 --> Estación\n"
		texto += "E-3,324801.87,2149732.37 --> Estación \n"
		texto += "\n"
		texto += "Puede ingresar la cantidad de estaciones que desee"

		ej.write(texto)
		ej.close()
		messagebox.showinfo(":)", "Ejemplo guardado")


#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

etiqueta = tk.Label(ventana, text = "Lineas de Conexión")
etiqueta.config(font = ("times new roman", 12), justify = "center")
etiqueta.place(x = 40, y = 10)

boton_cargar = tk.Button(ventana, text = "Cargar Archivo de Texto (.txt)", command = cargar)
boton_cargar.place(x = 10, y = 50)

boton_instructions = tk.Button(ventana, text = "Instrucciones", command = mostrar_instrucciones)
boton_instructions.place(x = 10, y = 80)

boton_exportar = tk.Button(ventana, text = "Exportar Tabla de Excel (.xlsx)", command = insertar_informacion)
boton_exportar.place(x = 10, y = 110)

boton_ejemplo = tk.Button(ventana, text = "Exportar Ejemplo (.txt)", command = exportar_ejemplo)
boton_ejemplo.place(x = 10, y = 140)

ventana.mainloop()
































































